"""
File utility functions for SFA system.
Handles file upload/download, CSV processing, Excel operations, and data validation.
"""

import os
import csv
import json
import zipfile
import shutil
import tempfile
from io import StringIO, BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional, Union, Generator, Tuple
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import mimetypes
import hashlib
import logging
from dataclasses import dataclass
from enum import Enum
import aiofiles
import asyncio
from concurrent.futures import ThreadPoolExecutor
import base64

logger = logging.getLogger(__name__)

class FileType(Enum):
    """Supported file types."""
    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"
    PDF = "pdf"
    IMAGE = "image"
    ZIP = "zip"
    TEXT = "txt"

@dataclass
class FileInfo:
    """File information container."""
    filename: str
    size: int
    mime_type: str
    file_type: FileType
    checksum: str
    upload_date: datetime
    path: Optional[str] = None
    metadata: Dict[str, Any] = None

class FileValidator:
    """File validation utilities."""
    
    # File size limits (in bytes)
    MAX_FILE_SIZES = {
        FileType.CSV: 50 * 1024 * 1024,      # 50 MB
        FileType.EXCEL: 100 * 1024 * 1024,   # 100 MB
        FileType.JSON: 10 * 1024 * 1024,     # 10 MB
        FileType.PDF: 25 * 1024 * 1024,      # 25 MB
        FileType.IMAGE: 5 * 1024 * 1024,     # 5 MB
        FileType.ZIP: 200 * 1024 * 1024,     # 200 MB
        FileType.TEXT: 1 * 1024 * 1024,      # 1 MB
    }
    
    # Allowed MIME types
    ALLOWED_MIME_TYPES = {
        FileType.CSV: ['text/csv', 'application/csv'],
        FileType.EXCEL: [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ],
        FileType.JSON: ['application/json', 'text/json'],
        FileType.PDF: ['application/pdf'],
        FileType.IMAGE: ['image/jpeg', 'image/png', 'image/gif', 'image/bmp'],
        FileType.ZIP: ['application/zip', 'application/x-zip-compressed'],
        FileType.TEXT: ['text/plain'],
    }
    
    @staticmethod
    def validate_file(file_path: str, expected_type: FileType = None) -> Tuple[bool, str]:
        """Validate file type, size, and integrity."""
        try:
            if not os.path.exists(file_path):
                return False, "File does not exist"
            
            # Check file size
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                return False, "File is empty"
            
            # Detect file type
            mime_type, _ = mimetypes.guess_type(file_path)
            detected_type = FileValidator._detect_file_type(file_path, mime_type)
            
            if expected_type and detected_type != expected_type:
                return False, f"Expected {expected_type.value}, got {detected_type.value}"
            
            # Check file size limits
            max_size = FileValidator.MAX_FILE_SIZES.get(detected_type, 50 * 1024 * 1024)
            if file_size > max_size:
                return False, f"File size ({file_size / 1024 / 1024:.1f} MB) exceeds limit ({max_size / 1024 / 1024:.1f} MB)"
            
            # Check MIME type
            allowed_mimes = FileValidator.ALLOWED_MIME_TYPES.get(detected_type, [])
            if mime_type and mime_type not in allowed_mimes:
                return False, f"MIME type {mime_type} not allowed for {detected_type.value}"
            
            return True, "File is valid"
            
        except Exception as e:
            return False, f"Validation error: {str(e)}"
    
    @staticmethod
    def _detect_file_type(file_path: str, mime_type: str = None) -> FileType:
        """Detect file type from extension and MIME type."""
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext in ['.csv']:
            return FileType.CSV
        elif file_ext in ['.xlsx', '.xls']:
            return FileType.EXCEL
        elif file_ext in ['.json']:
            return FileType.JSON
        elif file_ext in ['.pdf']:
            return FileType.PDF
        elif file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
            return FileType.IMAGE
        elif file_ext in ['.zip']:
            return FileType.ZIP
        elif file_ext in ['.txt']:
            return FileType.TEXT
        else:
            # Fallback to MIME type detection
            if mime_type:
                for file_type, mimes in FileValidator.ALLOWED_MIME_TYPES.items():
                    if mime_type in mimes:
                        return file_type
            
            return FileType.TEXT  # Default fallback

class FileManager:
    """Comprehensive file management for SFA system."""
    
    def __init__(self, upload_dir: str = "uploads", temp_dir: str = "temp"):
        self.upload_dir = Path(upload_dir)
        self.temp_dir = Path(temp_dir)
        
        # Create directories if they don't exist
        self.upload_dir.mkdir(exist_ok=True)
        self.temp_dir.mkdir(exist_ok=True)
        
        # File registry
        self.file_registry = {}
    
    def calculate_checksum(self, file_path: str) -> str:
        """Calculate MD5 checksum of file."""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def upload_file(self, file_content: bytes, filename: str, 
                   expected_type: FileType = None) -> Tuple[bool, Union[FileInfo, str]]:
        """Upload and validate file."""
        try:
                # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix)
            temp_file.write(file_content)
            temp_file.close()
            
            # Validate file
            is_valid, message = FileValidator.validate_file(temp_file.name, expected_type)
            if not is_valid:
                os.unlink(temp_file.name)
                return False, message
            
            # Calculate checksum
            checksum = self.calculate_checksum(temp_file.name)
            
            # Check for duplicates
            if checksum in self.file_registry:
                os.unlink(temp_file.name)
                return False, "File already exists"
            
            # Move to upload directory
            final_path = self.upload_dir / f"{checksum}_{filename}"
            shutil.move(temp_file.name, final_path)
            
            # Create file info
            file_info = FileInfo(
                filename=filename,
                size=len(file_content),
                mime_type=mimetypes.guess_type(filename)[0],
                file_type=FileValidator._detect_file_type(filename),
                checksum=checksum,
                upload_date=datetime.now(),
                path=str(final_path)
            )
            
            # Register file
            self.file_registry[checksum] = file_info
            
            logger.info(f"File uploaded successfully: {filename} ({checksum})")
            return True, file_info
            
        except Exception as e:
            logger.error(f"Upload failed: {str(e)}")
            return False, f"Upload failed: {str(e)}"

    def download_file(self, checksum: str) -> Tuple[bool, Union[bytes, str]]:
        """Download file by checksum."""
        try:
            if checksum not in self.file_registry:
                return False, "File not found"
            
            file_info = self.file_registry[checksum]
            
            if not os.path.exists(file_info.path):
                return False, "File path not found"
            
            with open(file_info.path, 'rb') as f:
                content = f.read()
            
            return True, content
            
        except Exception as e:
            logger.error(f"Download failed: {str(e)}")
            return False, f"Download failed: {str(e)}"

    def delete_file(self, checksum: str) -> Tuple[bool, str]:
        """Delete file by checksum."""
        try:
            if checksum not in self.file_registry:
                return False, "File not found"
            
            file_info = self.file_registry[checksum]
            
            if os.path.exists(file_info.path):
                os.unlink(file_info.path)
            
            del self.file_registry[checksum]
            
            logger.info(f"File deleted: {checksum}")
            return True, "File deleted successfully"
            
        except Exception as e:
            logger.error(f"Delete failed: {str(e)}")
            return False, f"Delete failed: {str(e)}"

    def list_files(self, file_type: FileType = None) -> List[FileInfo]:
        """List all uploaded files, optionally filtered by type."""
        files = list(self.file_registry.values())
        
        if file_type:
            files = [f for f in files if f.file_type == file_type]
        
        return sorted(files, key=lambda x: x.upload_date, reverse=True)

    async def upload_file_async(self, file_content: bytes, filename: str, 
                            expected_type: FileType = None) -> Tuple[bool, Union[FileInfo, str]]:
        """Async version of file upload."""
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor() as executor:
            return await loop.run_in_executor(
                executor, self.upload_file, file_content, filename, expected_type
            )

class CSVProcessor:
    """Enhanced CSV processing utilities for SFA system."""
    
    @staticmethod
    def read_csv(file_path: str, **kwargs) -> pd.DataFrame:
        """Read CSV with automatic encoding detection and error handling."""
        try:
            # Try common encodings
            encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding, **kwargs)
                    logger.info(f"CSV read successfully with {encoding} encoding")
                    return df
                except UnicodeDecodeError:
                    continue
            
            # If all fail, use error handling
            df = pd.read_csv(file_path, encoding='utf-8', errors='replace', **kwargs)
            logger.warning("CSV read with character replacement")
            return df
            
        except Exception as e:
            logger.error(f"Failed to read CSV: {str(e)}")
            raise

    @staticmethod
    def validate_csv_structure(df: pd.DataFrame, required_columns: List[str]) -> Tuple[bool, str]:
        """Validate CSV structure against required columns."""
        try:
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return False, f"Missing columns: {', '.join(missing_columns)}"
            
            if df.empty:
                return False, "CSV file is empty"
            
            return True, "CSV structure is valid"
            
        except Exception as e:
            return False, f"Validation error: {str(e)}"

    @staticmethod
    def clean_csv_data(df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize CSV data."""
        try:
            # Clean column names
            df.columns = df.columns.str.strip().str.replace(' ', '_')
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Strip whitespace from string columns
            string_columns = df.select_dtypes(include=['object']).columns
            df[string_columns] = df[string_columns].apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            
            # Convert date columns
            date_columns = [col for col in df.columns if 'date' in col.lower()]
            for col in date_columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            
            return df
            
        except Exception as e:
            logger.error(f"Data cleaning failed: {str(e)}")
            raise

    @staticmethod
    def process_sfa_gps_data(df: pd.DataFrame) -> pd.DataFrame:
        """Process SFA GPS data specifically."""
        try:
            # Validate required columns
            required_cols = ['DivisionCode', 'UserCode', 'UserName', 'Latitude', 'Longitude', 'RecievedDate']
            is_valid, message = CSVProcessor.validate_csv_structure(df, required_cols)
            if not is_valid:
                raise ValueError(message)
            
            # Clean data
            df = CSVProcessor.clean_csv_data(df)
            
            # Convert coordinates to numeric
            df['Latitude'] = pd.to_numeric(df['Latitude'], errors='coerce')
            df['Longitude'] = pd.to_numeric(df['Longitude'], errors='coerce')
            
            # Remove invalid coordinates
            df = df.dropna(subset=['Latitude', 'Longitude'])
            df = df[(df['Latitude'].between(-90, 90)) & (df['Longitude'].between(-180, 180))]
            
            # Sort by user and date
            df = df.sort_values(['UserCode', 'RecievedDate'])
            
            # Add calculated fields
            df['Speed'] = CSVProcessor._calculate_speed(df)
            df['Distance'] = CSVProcessor._calculate_distance(df)
            
            return df
            
        except Exception as e:
            logger.error(f"GPS data processing failed: {str(e)}")
            raise

    @staticmethod
    def _calculate_speed(df: pd.DataFrame) -> pd.Series:
        """Calculate speed between GPS points."""
        speeds = []
        for user in df['UserCode'].unique():
            user_data = df[df['UserCode'] == user].copy()
            user_speeds = [0]  # First point has 0 speed
            
            for i in range(1, len(user_data)):
                prev_row = user_data.iloc[i-1]
                curr_row = user_data.iloc[i]
                
                # Calculate distance (km)
                distance = CSVProcessor._haversine_distance(
                    prev_row['Latitude'], prev_row['Longitude'],
                    curr_row['Latitude'], curr_row['Longitude']
                )
                
                # Calculate time difference (hours)
                time_diff = (curr_row['RecievedDate'] - prev_row['RecievedDate']).total_seconds() / 3600
                
                # Calculate speed (km/h)
                speed = distance / time_diff if time_diff > 0 else 0
                user_speeds.append(min(speed, 200))  # Cap at 200 km/h to filter outliers
            
            speeds.extend(user_speeds)
        
        return pd.Series(speeds, index=df.index)

    @staticmethod
    def _calculate_distance(df: pd.DataFrame) -> pd.Series:
        """Calculate cumulative distance traveled."""
        distances = []
        for user in df['UserCode'].unique():
            user_data = df[df['UserCode'] == user].copy()
            user_distances = [0]  # First point has 0 cumulative distance
            
            cumulative_distance = 0
            for i in range(1, len(user_data)):
                prev_row = user_data.iloc[i-1]
                curr_row = user_data.iloc[i]
                
                distance = CSVProcessor._haversine_distance(
                    prev_row['Latitude'], prev_row['Longitude'],
                    curr_row['Latitude'], curr_row['Longitude']
                )
                
                cumulative_distance += distance
                user_distances.append(cumulative_distance)
            
            distances.extend(user_distances)
        
        return pd.Series(distances, index=df.index)

    @staticmethod
    def _haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """Calculate haversine distance between two points in kilometers."""
        from math import radians, cos, sin, asin, sqrt
        
        # Convert to radians
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        
        # Haversine formula
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * asin(sqrt(a))
        r = 6371  # Earth's radius in kilometers
        
        return c * r

class ExcelProcessor:
    """Advanced Excel processing for SFA system."""
    
    @staticmethod
    def read_excel_all_sheets(file_path: str) -> Dict[str, pd.DataFrame]:
        """Read all sheets from Excel file."""
        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                df = CSVProcessor.clean_csv_data(df)
                sheets[sheet_name] = df
                logger.info(f"Sheet '{sheet_name}' loaded with {len(df)} rows")
            
            return sheets
            
        except Exception as e:
            logger.error(f"Failed to read Excel file: {str(e)}")
            raise

    @staticmethod
    def process_sfa_orders(file_path: str) -> pd.DataFrame:
        """Process SFA Orders Excel file with multiple sheets."""
        try:
            sheets = ExcelProcessor.read_excel_all_sheets(file_path)
            
            # Combine all monthly sheets
            all_orders = []
            for sheet_name, df in sheets.items():
                if not df.empty:
                    df['Month'] = sheet_name
                    all_orders.append(df)
            
            if not all_orders:
                raise ValueError("No valid data found in Excel sheets")
            
            combined_df = pd.concat(all_orders, ignore_index=True)
            
            # Validate required columns
            required_cols = ['Code', 'Date', 'DistributorCode', 'UserCode', 'FinalValue']
            is_valid, message = CSVProcessor.validate_csv_structure(combined_df, required_cols)
            if not is_valid:
                raise ValueError(message)
            
            # Process data
            combined_df['FinalValue'] = pd.to_numeric(combined_df['FinalValue'], errors='coerce')
            combined_df = combined_df.dropna(subset=['FinalValue'])
            
            # Add calculated fields
            combined_df['OrderValue'] = combined_df['FinalValue']
            combined_df['OrderMonth'] = pd.to_datetime(combined_df['Date']).dt.to_period('M')
            combined_df['OrderYear'] = pd.to_datetime(combined_df['Date']).dt.year
            
            return combined_df
            
        except Exception as e:
            logger.error(f"SFA Orders processing failed: {str(e)}")
            raise

    @staticmethod
    def create_report_excel(data: Dict[str, pd.DataFrame], output_path: str):
        """Create formatted Excel report with multiple sheets."""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get worksheet for formatting
                    worksheet = writer.sheets[sheet_name]
                    
                    # Style headers
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Excel report created: {output_path}")
            
        except Exception as e:
            logger.error(f"Excel report creation failed: {str(e)}")
            raise

class DataExporter:
    """Data export utilities for various formats."""
    
    @staticmethod
    def export_to_json(data: Union[pd.DataFrame, Dict], output_path: str):
        """Export data to JSON format."""
        try:
            if isinstance(data, pd.DataFrame):
                data.to_json(output_path, orient='records', date_format='iso', indent=2)
            else:
                with open(output_path, 'w') as f:
                    json.dump(data, f, indent=2, default=str)
            
            logger.info(f"Data exported to JSON: {output_path}")
            
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            raise

    @staticmethod
    def export_to_csv(df: pd.DataFrame, output_path: str, **kwargs):
        """Export DataFrame to CSV with encoding handling."""
        try:
            df.to_csv(output_path, index=False, encoding='utf-8-sig', **kwargs)
            logger.info(f"Data exported to CSV: {output_path}")
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise

    @staticmethod
    def create_zip_archive(files: List[str], output_path: str):
        """Create ZIP archive from multiple files."""
        try:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in files:
                    if os.path.exists(file_path):
                        arcname = os.path.basename(file_path)
                        zipf.write(file_path, arcname)
                        logger.info(f"Added to archive: {arcname}")
            
            logger.info(f"ZIP archive created: {output_path}")
            
        except Exception as e:
            logger.error(f"ZIP creation failed: {str(e)}")
            raise

# Utility functions for specific SFA data processing
def process_customer_data(file_path: str) -> pd.DataFrame:
    """Process customer data file."""
    try:
        if file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path)
        else:
            df = CSVProcessor.read_csv(file_path)
        
        df = CSVProcessor.clean_csv_data(df)
        
        # Validate customer data structure
        if 'City' in df.columns:
            df['City'] = df['City'].str.title()
        
        return df
        
    except Exception as e:
        logger.error(f"Customer data processing failed: {str(e)}")
        raise

def merge_customer_location_data(customer_df: pd.DataFrame, location_df: pd.DataFrame) -> pd.DataFrame:
    """Merge customer data with location coordinates."""
    try:
        # Ensure both DataFrames have common identifier
        merge_key = None
        for key in ['No.', 'CustomerID', 'ID', 'Code']:
            if key in customer_df.columns and key in location_df.columns:
                merge_key = key
                break
        
        if not merge_key:
            raise ValueError("No common identifier found for merging")
        
        merged_df = customer_df.merge(location_df, on=merge_key, how='left')
        
        # Validate coordinates
        if 'Latitude' in merged_df.columns and 'Longitude' in merged_df.columns:
            merged_df['Latitude'] = pd.to_numeric(merged_df['Latitude'], errors='coerce')
            merged_df['Longitude'] = pd.to_numeric(merged_df['Longitude'], errors='coerce')
        
        return merged_df
        
    except Exception as e:
        logger.error(f"Data merging failed: {str(e)}")
        raise

# Initialize global file manager
file_manager = FileManager()

# Export main classes and functions
__all__ = [
    'FileType', 'FileInfo', 'FileValidator', 'FileManager',
    'CSVProcessor', 'ExcelProcessor', 'DataExporter',
    'process_customer_data', 'merge_customer_location_data', 'file_manager'
]